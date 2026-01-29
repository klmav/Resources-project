from __future__ import annotations

import datetime as dt
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl

from ..sheets.client import SheetValues


@dataclass(frozen=True)
class XlsxParseHints:
    """
    Подсказки для парсинга.
    В MVP стараемся авто-распознать шапку, но оставляем ручные override.
    """

    header_row: Optional[int] = None
    max_cols: Optional[int] = None
    include_row_above_header: bool = True


def _is_month_header_cell(v: object) -> bool:
    return isinstance(v, dt.datetime) or isinstance(v, dt.date)


def _guess_header_row(ws, scan_rows: int = 30) -> int:
    """
    Ищем строку, похожую на шапку: содержит 'A/D' и хотя бы одну дату-месяц.
    """
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        first = ws.cell(r, 1).value
        if isinstance(first, str) and first.strip().upper() == "A/D":
            # look for any date-like header cell in row
            for c in range(1, min(ws.max_column, 200) + 1):
                if _is_month_header_cell(ws.cell(r, c).value):
                    return r
    # fallback to 1
    return 1


def read_xlsx_as_sheet_values(
    path: str | Path,
    sheet_name: Optional[str] = None,
    hints: Optional[XlsxParseHints] = None,
) -> SheetValues:
    p = Path(path)
    wb = openpyxl.load_workbook(p, data_only=True)

    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    hints = hints or XlsxParseHints()

    header_row = hints.header_row or _guess_header_row(ws)

    read_from_row = header_row
    if hints.include_row_above_header and header_row > 1:
        read_from_row = header_row - 1

    # Determine how many columns to read: until trailing empty cells in header row,
    # unless explicitly overridden.
    max_cols = hints.max_cols
    if max_cols is None:
        max_cols = 0
        for c in range(1, ws.max_column + 1):
            if ws.cell(header_row, c).value is not None:
                max_cols = c
        max_cols = max(max_cols, 1)

    # Read from header row down to last used row.
    values: list[list[str]] = []
    for r in range(read_from_row, ws.max_row + 1):
        row: list[str] = []
        empty = True
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            if v is None:
                row.append("")
                continue
            empty = False
            if isinstance(v, (dt.datetime, dt.date)):
                row.append(v.strftime("%Y-%m-%d"))
            else:
                row.append(str(v).strip() if isinstance(v, str) else str(v))
        # stop if we reached long tail of empty rows
        if empty and r > header_row + 5:
            break
        values.append(row)

    return SheetValues(values=values)

