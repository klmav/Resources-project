from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill

from .models import FixPlan


@dataclass(frozen=True)
class ApplyResult:
    output_path: Path
    applied_count: int


def apply_fix_plan_to_xlsx_copy(
    *,
    input_path: str | Path,
    plan: FixPlan,
    output_path: Optional[str | Path] = None,
) -> ApplyResult:
    src = Path(input_path)
    if output_path is None:
        out = src.with_name(src.stem + ".fixed" + src.suffix)
    else:
        out = Path(output_path)

    # Copy bytes first (keep originals intact)
    out.write_bytes(src.read_bytes())

    wb = openpyxl.load_workbook(out)
    ws = (
        wb[plan.sheet_name]
        if plan.sheet_name in wb.sheetnames
        else wb[wb.sheetnames[0]]
    )

    applied = 0
    for upd in plan.updates:
        if upd.sheet_name != plan.sheet_name:
            continue
        cell = ws.cell(upd.cell.row, upd.cell.col)
        if upd.kind == "set_value":
            cell.value = upd.new_value
        elif upd.kind == "highlight":
            rgb = upd.fill_rgb or "FFFFC7CE"  # light red default
            cell.fill = PatternFill(patternType="solid", fgColor=rgb)
        else:
            # unknown update kind - skip safely
            continue
        applied += 1

    wb.save(out)
    return ApplyResult(output_path=out, applied_count=applied)
